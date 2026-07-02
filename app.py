from datetime import datetime
from io import BytesIO
from copy import copy

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

TEMPLATE_FILE = "tdr_template.xlsx"

st.set_page_config(page_title="TDR Generator", page_icon="⚓", layout="wide")

st.markdown("""
# ⚓ TDR Generator
ALPORT Terminal Departure Report oluşturma sistemi
""")


def write_value(ws, cell, value):
    """Boş olmayan değeri ilgili hücreye yazar."""
    if value is not None and value != "":
        ws[cell] = value


def clear_table(ws, start_row, max_row, columns):
    """Belirtilen tablo alanındaki eski verileri temizler."""
    for row in range(start_row, max_row + 1):
        for col in columns:
            ws.cell(row=row, column=col).value = None


def copy_row_style(ws, source_row, target_row, max_col):
    """Şablondaki satır formatını yeni satıra kopyalar."""
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


def fill_basic_info(wb, data):
    """Temel gemi bilgilerini MOVE COUNT sayfasına yazar."""
    ws = wb["MOVE COUNT"]

    # İlk sürümde ana hücreler. Şablon detayları netleşince buraya yeni hücreler eklenebilir.
    write_value(ws, "E3", data.get("vessel_name"))
    write_value(ws, "E4", data.get("voyage_no"))
    write_value(ws, "E5", data.get("flag"))


def fill_container_sheet(wb, sheet_name, df, start_row=3):
    """Yüklenen Excel listesini ilgili TDR sayfasına yazar."""
    if df is None or df.empty:
        return

    if sheet_name not in wb.sheetnames:
        return

    ws = wb[sheet_name]
    clear_table(ws, start_row, ws.max_row, range(1, ws.max_column + 1))

    df = df.fillna("")

    for idx, row in df.iterrows():
        excel_row = start_row + idx
        copy_row_style(ws, start_row, excel_row, ws.max_column)
        ws.cell(excel_row, 1).value = idx + 1

        for col_idx, value in enumerate(row.tolist(), start=2):
            ws.cell(excel_row, col_idx).value = value


def read_uploaded_excel(uploaded_file):
    """Kullanıcının yüklediği Excel dosyasını okur."""
    if uploaded_file is None:
        return None

    try:
        return pd.read_excel(uploaded_file)
    except Exception as exc:
        st.error(f"Excel okunamadı: {exc}")
        return None


def generate_tdr_excel(data, discharge_df=None, loading_df=None, reefer_df=None, imdg_df=None, damage_df=None):
    """Şablonu doldurup indirilebilir Excel verisi döndürür."""
    wb = load_workbook(TEMPLATE_FILE)

    fill_basic_info(wb, data)
    fill_container_sheet(wb, "DISCHARGING", discharge_df)
    fill_container_sheet(wb, "LOADING", loading_df)
    fill_container_sheet(wb, "REEFER", reefer_df)
    fill_container_sheet(wb, "IMDG", imdg_df)
    fill_container_sheet(wb, "DAMAGELIST", damage_df)

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


def safe_filename(text):
    """Dosya adı için güvenli metin üretir."""
    value = str(text or "").strip().upper().replace(" ", "_")
    return value if value else "UNKNOWN"


with st.sidebar:
    st.header("Menü")
    st.info("Bu sürüm klasör kullanmaz. app.py ve tdr_template.xlsx aynı yerde olmalıdır.")

st.subheader("1) Gemi Bilgileri")
col1, col2, col3 = st.columns(3)

with col1:
    vessel_name = st.text_input("Vessel Name", value="MARTI SPIRIT")
with col2:
    voyage_no = st.text_input("Voyage No", value="625S")
with col3:
    flag = st.text_input("Flag", value="PORTUGAL")

st.subheader("2) Liste Yükleme")
st.caption(
    "Kolon sırası şablonla aynı olmalı. Örnek DISCHARGING: "
    "Container No, Size, Type, Iso Code, Discharge Time-Date, Gross Weight, F / E"
)

c1, c2 = st.columns(2)
with c1:
    discharge_file = st.file_uploader("Discharging Excel", type=["xlsx", "xls"], key="discharge")
    reefer_file = st.file_uploader("Reefer Excel", type=["xlsx", "xls"], key="reefer")
    damage_file = st.file_uploader("Damage List Excel", type=["xlsx", "xls"], key="damage")
with c2:
    loading_file = st.file_uploader("Loading Excel", type=["xlsx", "xls"], key="loading")
    imdg_file = st.file_uploader("IMDG Excel", type=["xlsx", "xls"], key="imdg")

st.subheader("3) Rapor Oluştur")

if st.button("TDR Excel Oluştur", type="primary"):
    data = {
        "vessel_name": vessel_name,
        "voyage_no": voyage_no,
        "flag": flag,
    }

    discharge_df = read_uploaded_excel(discharge_file)
    loading_df = read_uploaded_excel(loading_file)
    reefer_df = read_uploaded_excel(reefer_file)
    imdg_df = read_uploaded_excel(imdg_file)
    damage_df = read_uploaded_excel(damage_file)

    try:
        excel_buffer = generate_tdr_excel(
            data=data,
            discharge_df=discharge_df,
            loading_df=loading_df,
            reefer_df=reefer_df,
            imdg_df=imdg_df,
            damage_df=damage_df,
        )

        filename = f"TDR-{safe_filename(vessel_name)}-{safe_filename(voyage_no)}.xlsx"

        st.success("TDR başarıyla oluşturuldu.")
        st.download_button(
            label="Excel Dosyasını İndir",
            data=excel_buffer,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except FileNotFoundError:
        st.error("tdr_template.xlsx bulunamadı. Bu dosya app.py ile aynı yerde olmalı.")
    except Exception as exc:
        st.error(f"TDR oluşturulurken hata oluştu: {exc}")

st.divider()
st.subheader("Excel Kolon Formatları")
st.markdown("""
**DISCHARGING:** Container No, Size, Type, Iso Code, Discharge Time-Date, Gross Weight, F / E  
**LOADING:** Container No, Size, Type, Iso Code, Load Time-Date, Last Location, Gross Weight, F / E  
**REEFER:** Container No, Size, Type, Iso Code, Gross Weight, F / E, Seal  
**IMDG:** Container No, Size, Type, Iso Code, Gross Weight, F / E, Seal  
**DAMAGELIST:** Container No, Agent, Liner, Operator, Container Size, Container Group, Container Type
""")
